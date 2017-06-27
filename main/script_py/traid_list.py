from openpyxl import load_workbook
from openpyxl import Workbook
import os
import re

def main(file_name):
    wb = load_workbook(file_name)
    ws=wb[wb.get_sheet_names()[0]]


    test_str="трейд-ин"
    a1="Рекомендованная розничная цена"
    a2="Общая цена"
    a3="Таким образом, итоговая стоимость"
    discount111="от вышеуказанной общей цены"
    for row in list(ws.iter_rows(range_string="A1:E90")):
        for cell in row:
            if cell.value==None:
                continue
            else:
                if a1 in str(cell.value):
                    recommended_price=cell.value
                elif a2 in str(cell.value):
                    total=cell.value
                elif a3 in str(cell.value):
                    itog=cell.value
                elif discount111 in str(cell.value):
                    discount=cell.value
    recommended_price=re.findall(r'\d?\s?\d\d\d+\s\d\d\d+',recommended_price)[-1]
    total=re.findall(r'\d?\s?\d\d\d+\s\d\d\d+',total)[-1]
    discount=re.findall(r'\b\d\d\s\d\d\d\b\s',discount)[-1]
    vin_new_car=ws['E12'].value
    vin_old_car=ws['E27'].value
    week=25
# print("recommended price:",recommended_price)#H20
# print("total",total)#H21
# print("discount",discount)#H22
# print("vin_new_car",vin_new_car)#D13
# print("vin_old_car",vin_old_car)#D15
# print("week",week)#H1

    wd_save=load_workbook("list traid.xlsx")
    ws_save=wd_save[wd_save.get_sheet_names()[0]]
    ws_save["D13"].value=vin_new_car
    ws_save["D15"].value=vin_old_car
    ws_save["H1"].value=week
    ws_save["H20"].value=recommended_price
    ws_save["H21"].value=total
    ws_save["H22"].value=discount

    wd_save.save('list traid'+file_name+'.xlsx')



file_list=list()
for i in os.listdir(path="."):
    if i.endswith(".xlsx"):
        if i!="list traid.xlsx" or not i.startswith("list traid"):
            file_list.append(i)

if  __name__ ==  "__main__" :
    for i in file_list:
        print(i)
        main(i)

def clear():
    a=os.listdir(path=".")
    for i in a:
        if i.endswith(".xlsx"):
            if (i!='qw.py'
            and i!='list traid.xlsx'):
                os.system("rm \'{a1}'".format(a1=i))
                print("del {a1}".format(a1=i))
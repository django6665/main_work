# from openpyxl import load_workbook
# from openpyxl import Workbook
# import os
# import re
# def main(file_name,path):
#     wb = load_workbook(file_name)
#     ws=wb[wb.get_sheet_names()[0]]
#     try:
#         wb_save = load_workbook(path+'/'+'main.xlsx')
#     except:
#         wb_save = Workbook()
#     ws_save = wb_save.active
# #######################ws_save first_row
#     ws_save["A1"].value=r"№ п/п"
#     ws_save["B1"].value="BIR-код дилера"
#     ws_save["C1"].value="Полное юридическое наименование дилера"
#     ws_save["D1"].value="Механизм акции (утилизация/трейд-ин)"
#     ws_save["E1"].value="Дата продажи нового автомобиля (в соответствие с Актом приема-передачи)"
#     ws_save["F1"].value="VIN нового автомобиля"
#     ws_save["G1"].value="Марка нового автомобиля"
#     ws_save["H1"].value="Модель нового автомобиля"
#     ws_save["I1"].value="Год выпуска нового автомобиля"
#     ws_save["J1"].value="Общая цена нового Автомобиля по Договору с учетом дополнительных опций,стоимости доставки, кредитной субсидии и скидки дилера в руб. с НДС"
#     ws_save["K1"].value="Размер скидки по Программе обновления автомобилей в руб. с НДС"
#     ws_save["L1"].value="VIN сдаваемого по программе автомобиля"
#     ws_save["M1"].value="Серия и № ПТС (в случае отсутствия VIN)"
#     ws_save["N1"].value="Серия и № св-ва о регистрации ТС (в случае отсутствия VIN)"
#     ws_save["O1"].value="Марка сдаваемого по программе автомобиля"
#     ws_save["P1"].value="Модель сдаваемого по программе автомобиля"
#     ws_save["Q1"].value="Возраст сдаваемого по программе автомобиля (полных лет) = текущий год – год производства автомобиля по ПТС"
#     ws_save["R1"].value="Покупатель (физ. лицо / юр. лицо)"
#     ws_save["S1"].value="Пробег, км"
#     ws_save["T1"].value="Количество владельцев (по ПТС)"
#     ws_save["U1"].value="Утилизатор"
#     ws_save["V1"].value="Дополнительно реализуемая мера (субсидированный кредит Renault)"
# #######################it is trade_in or uti
#     test_str="трейд-ин"
#     a1="Рекомендованная розничная цена"
#     a2="Общая цена"
#     a3="Таким образом, итоговая стоимость"
#     discount111="от вышеуказанной общей цены"
#     for row in list(ws.iter_rows(range_string="A1:E90")):
#         for cell in row:
#             if cell.value==None:
#                 continue
#             else:
#                 if a1 in str(cell.value):
#                     recommended_price=cell.value
#                 elif a2 in str(cell.value):
#                     total=cell.value
#                 elif a3 in str(cell.value):
#                     itog=cell.value
#                 elif discount111 in str(cell.value):
#                     discount=cell.value
#     recommended_price=re.findall(r'\d?\s?\d\d\d+\s\d\d\d+',recommended_price)[-1]
#     total=re.findall(r'\d?\s?\d\d\d+\s\d\d\d+',total)[-1]
#     discount=re.findall(r'\b\d\d\s\d\d\d\b\s',discount)[-1]

# ###################trade-in
#     if ws["B23"].value!=None:
#             B_ws="64306771"
#             C_ws="ООО \"Автосалон Престиж\""
#             D_ws="трейд-ин"
#             #E_ws= not in contract
#             F_ws=ws["E12"].value
#             G_ws=ws["E10"].value
#             H_ws=ws["E11"].value
#             I_ws=ws["E17"].value
#             J_ws=total ###total cost
#             K_ws=discount###discount
#             L_ws=ws["E27"].value
#             #######################start:old-new car
#             frodo=ws["E25"].value
#             www=re.search(r"\b\W",frodo)
#             qq=www.span()[0]
#             aaa0=frodo[:qq]
#             aaa1=frodo[(qq+1):]
#             aaa0=re.sub(r"\W\s"," ",aaa0,count=1)
#             aaa1=re.sub(r"\W\s"," ",aaa1,count=1)
#             ###################end: old-new car
#             O_ws=aaa0#old model car
#             P_ws=aaa1#old version car
#             Q_ws=ws["E37"].value
#             R_ws="физ. лицо"
#             S_ws=ws["E35"].value
#             T_ws=ws["E38"].value
#             U_ws="-"
# ###################utilization
#     else:
#         B_ws="64306771"
#         C_ws="ООО \"Автосалон Престиж\""
#         D_ws="утилизация"
#         F_ws=ws["E13"].value
#         G_ws=ws["E11"].value
#         H_ws=ws["E12"].value
#         I_ws=ws["E18"].value
#         J_ws=total
#         K_ws=discount
#         L_ws=ws["E28"].value
#         #######################start:old-new car
#         frodo=ws["E26"].value
#         www=re.search(r"\b\W",frodo)
#         qq=www.span()[0]
#         aaa0=frodo[:qq]
#         aaa1=frodo[(qq+1):]
#         aaa0=re.sub(r"\W\s"," ",aaa0,count=1)
#         aaa1=re.sub(r"\W\s"," ",aaa1,count=1)
#         ###################end: old-new car
#         O_ws=aaa0
#         P_ws=aaa1
#         Q_ws=ws["E38"].value
#         R_ws="физ. лицо"
#         S_ws=ws["E36"].value
#         T_ws=ws["E39"].value
#         U_ws="-"
# #################save
#     ws_save.append({
#       "B":B_ws,
#       "C":C_ws,
#       "D":D_ws,
#       "F":F_ws,
#       "G":G_ws,
#       "H":H_ws,
#       "I":I_ws,
#       "J":J_ws,
#       "K":K_ws,
#       "L":L_ws,
#       "O":O_ws,
#       "P":P_ws,
#       "Q":Q_ws,
#       "R":R_ws,
#       "S":S_ws,
#       "T":T_ws,
#       "U":U_ws,
#     })
#     wb_save.save(path+'/'+'main.xlsx')
# ##########run
# # file_list=list()
# # for i in os.listdir(path="."):
# #     if i.endswith(".xlsx"):
# #         if i!="main.xlsx":
# #             file_list.append(i)

# # for i in file_list:
# #     print(i)
# #     main(i)



from openpyxl import load_workbook
from openpyxl import Workbook
import re
def main(file_name,path,week):
    wb = load_workbook(file_name)
    ws=wb[wb.get_sheet_names()[0]]
    try:
        wb_save = load_workbook(path+'/'+'main.xlsx')
    except:
        wb_save = Workbook()
    ws_save = wb_save.active
    #ws_save first_row
    ws_save["A1"].value=r"№ п/п"
    ws_save["B1"].value="BIR-код дилера"
    ws_save["C1"].value="Полное юридическое наименование дилера"
    ws_save["D1"].value="Механизм акции (утилизация/трейд-ин)"
    ws_save["E1"].value="Дата продажи нового автомобиля (в соответствие с Актом приема-передачи)"
    ws_save["F1"].value="VIN нового автомобиля"
    ws_save["G1"].value="Марка нового автомобиля"
    ws_save["H1"].value="Модель нового автомобиля"
    ws_save["I1"].value="Год выпуска нового автомобиля"
    ws_save["J1"].value="Общая цена нового Автомобиля по Договору с учетом дополнительных опций,стоимости доставки, кредитной субсидии и скидки дилера в руб. с НДС"
    ws_save["K1"].value="Размер скидки по Программе обновления автомобилей в руб. с НДС"
    ws_save["L1"].value="VIN сдаваемого по программе автомобиля"
    ws_save["M1"].value="Серия и № ПТС (в случае отсутствия VIN)"
    ws_save["N1"].value="Серия и № св-ва о регистрации ТС (в случае отсутствия VIN)"
    ws_save["O1"].value="Марка сдаваемого по программе автомобиля"
    ws_save["P1"].value="Модель сдаваемого по программе автомобиля"
    ws_save["Q1"].value="Возраст сдаваемого по программе автомобиля (полных лет) = текущий год – год производства автомобиля по ПТС"
    ws_save["R1"].value="Покупатель (физ. лицо / юр. лицо)"
    ws_save["S1"].value="Пробег, км"
    ws_save["T1"].value="Количество владельцев (по ПТС)"
    ws_save["U1"].value="Утилизатор"
    ws_save["V1"].value="Дополнительно реализуемая мера (субсидированный кредит Renault)"
#######################it is trade_in or uti
    test_str="трейд-ин"
    a1="Рекомендованная розничная цена"
    a2="Общая цена"
    a3="Таким образом, итоговая стоимость"
    discount111="от вышеуказанной общей цены"
    for row in list(ws.iter_rows(range_string="A1:E90")):
        for cell in row:
            if cell.value==None:
                continue
            else:
                if a1 in str(cell.value):
                    recommended_price=cell.value
                elif a2 in str(cell.value):
                    total=cell.value
                elif a3 in str(cell.value):
                    itog=cell.value
                elif discount111 in str(cell.value):
                    discount=cell.value
    recommended_price=re.findall(r'\d?\s?\d\d\d+\s\d\d\d+',recommended_price)[-1]
    total=re.findall(r'\d?\s?\d\d\d+\s\d\d\d+',total)[-1]
    discount=re.findall(r'\b\d\d\s\d\d\d\b\s',discount)[-1]

###################trade-in
    if ws["B23"].value!=None:
            action="traid"
            B_ws="64306771"
            C_ws="ООО \"Автосалон Престиж\""
            D_ws="трейд-ин"
            F_ws=ws["E12"].value
            G_ws=ws["E10"].value
            H_ws=ws["E11"].value
            I_ws=ws["E17"].value
            J_ws=total ###total cost
            K_ws=discount###discount
            L_ws=ws["E27"].value
            #######################start:old-new car
            frodo=ws["E25"].value
            www=re.search(r"\b\W",frodo)
            qq=www.span()[0]
            aaa0=frodo[:qq]
            aaa1=frodo[(qq+1):]
            aaa0=re.sub(r"\W\s"," ",aaa0,count=1)
            aaa1=re.sub(r"\W\s"," ",aaa1,count=1)
            ###################end: old-new car
            O_ws=aaa0#old model car
            P_ws=aaa1#old version car
            Q_ws=ws["E37"].value
            R_ws="физ. лицо"
            S_ws=ws["E35"].value
            T_ws=ws["E38"].value
            U_ws="-"
###################utilization
    else:
        action="no_traid"
        B_ws="64306771"
        C_ws="ООО \"Автосалон Престиж\""
        D_ws="утилизация"
        F_ws=ws["E13"].value
        G_ws=ws["E11"].value
        H_ws=ws["E12"].value
        I_ws=ws["E18"].value
        J_ws=total
        K_ws=discount
        L_ws=ws["E28"].value
        #######################start:old-new car
        frodo=ws["E26"].value
        www=re.search(r"\b\W",frodo)
        qq=www.span()[0]
        aaa0=frodo[:qq]
        aaa1=frodo[(qq+1):]
        aaa0=re.sub(r"\W\s"," ",aaa0,count=1)
        aaa1=re.sub(r"\W\s"," ",aaa1,count=1)
        ###################end: old-new car
        O_ws=aaa0
        P_ws=aaa1
        Q_ws=ws["E38"].value
        R_ws="физ. лицо"
        S_ws=ws["E36"].value
        T_ws=ws["E39"].value
        U_ws="-"
#################save
    ws_save.append({
      "B":B_ws,
      "C":C_ws,
      "D":D_ws,
      "F":F_ws,
      "G":G_ws,
      "H":H_ws,
      "I":I_ws,
      "J":J_ws,
      "K":K_ws,
      "L":L_ws,
      "O":O_ws,
      "P":P_ws,
      "Q":Q_ws,
      "R":R_ws,
      "S":S_ws,
      "T":T_ws,
      "U":U_ws,
    })
    wb_save.save(path+'/'+'main.xlsx')




    # if action=="traid":
    #     file_name_list_xlsx="Копия Сопроводительный лист ТРЕЙД-ИН новый.xlsx"
    # else:
    #     file_name_list_xlsx="Копия Сопроводительный лист УТИЛИЗАЦИЯ новый.xlsx"
    # path_to_list="/home/django6665/main/script_py"
    # list_xlsx=load_workbook('{path_to_list}/{file_name_list_xlsx}'.format(file_name_list_xlsx=file_name_list_xlsx,
    #                                                                 path_to_list=path_to_list))
    if action=="traid":
        list_xlsx=load_workbook("/home/django6665/main/script_py/Копия Сопроводительный лист ТРЕЙД-ИН новый.xlsx")
        ws_list_xlsx = list_xlsx.active
        ws_list_xlsx["D13"]=ws["E12"].value #vin new car
        ws_list_xlsx["D15"]=ws["E27"].value #VIN OLD CAR
        ws_list_xlsx["H1"]=week####nomber week

        ws_list_xlsx["H20"]=recommended_price #Рекомендованная розничная цена версии по тарифу
        ws_list_xlsx["H21"]=total #Цена с учетом доп.опций, акс-ов, дост-ки, минус скидка дилера и/или субсидия
        ws_list_xlsx["H122"]=discount #Скидка по программе

        list_xlsx.save(path+'/'+"list"+str(ws["E12"].value)+'.xlsx')
    else:
        list_xlsx=load_workbook("/home/django6665/main/script_py/Копия Сопроводительный лист УТИЛИЗАЦИЯ новый.xlsx")
        ws_list_xlsx = list_xlsx.active
        ws_list_xlsx["D13"]=ws["E13"].value #vin new car
        ws_list_xlsx["D15"]=ws["E28"].value #VIN OLD CAR
        ws_list_xlsx["H1"]=week####nomber week

        ws_list_xlsx["H20"]=recommended_price #Рекомендованная розничная цена версии по тарифу
        ws_list_xlsx["H21"]=total #Цена с учетом доп.опций, акс-ов, дост-ки, минус скидка дилера и/или субсидия
        ws_list_xlsx["H122"]=discount #Скидка по программе

        list_xlsx.save(path+'/'+"list"+str(ws["E28"].value)+'.xlsx')

